import pytest
from selenium import webdriver
import openpyxl
from datetime import datetime
from login_page import LoginPage  
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Read Excel data for the tests
def get_test_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

# Write test result back to Excel
def write_test_result(test_id, result, file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == test_id:
            sheet.cell(row=row[0].row, column=5).value = datetime.now().strftime('%H:%M:%S')
            sheet.cell(row=row[0].row, column=4).value = datetime.now().strftime('%Y-%m-%d')
            sheet.cell(row=row[0].row, column=7).value = result
            break
    workbook.save(file_path)

@pytest.mark.parametrize("test_id,username,password,date,time_of_test,name_of_tester,test_result", get_test_data("login_test_data.xlsx"))
def test_login(test_id, username, password, date, time_of_test, name_of_tester, test_result):
    driver = webdriver.Chrome()  # Assumes chromedriver is in your PATH
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    
    login_page = LoginPage(driver)
    login_page.login(username, password)
    
    # Check if login was successful (using an element unique to the landing page)
    try:
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//p[contains(text(),'Dashboard')]")))
        result = "Passed"
    except:
        result = "Failed"
    
    # Write result back to Excel
    write_test_result(test_id, result, "login_test_data.xlsx")
    
    driver.quit()
